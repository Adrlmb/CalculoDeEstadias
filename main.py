import pdfplumber
from openpyxl import load_workbook

pdf = pdfplumber.open('C:/Users/adrie/Downloads/TICKET 1.pdf')
page = pdf.pages[0]
text = page.extract_text()

wb = load_workbook('estadia\Cálculo estadia.xlsx')  # Carrega o arquivo existente
planilha = wb.active  # Seleciona a planilha ativa


def dataDeSaida():
    data = text.split('\n')[4].split(' ')[2].replace('.', '/')
    hora = text.split('\n')[4].split(' ')[3].split(':')[0]
    minutos = text.split('\n')[4].split(' ')[3].split(':')[1]

    return data + ' ' + hora + ':' + minutos


planilha['F3'] = text.split('\n')[10].split('-')[1]  # Transportador
planilha['C4'] = int(text.split('\n')[7].split(':')[1])  # NF
planilha['C5'] = text.split('\n')[9].split('-')[1]  # Produto, impor condição se for Rocha!!
planilha['F12'] = text.split('\n')[5].split(": ")[1]  # Peso NF com vírgula
planilha['B15'] = dataDeSaida()  # Data e Hora da saída no formato DD/MM/AA HH:MM

planilha['B9'] = 'INPUT'  # Data e Hora de chegada no formato DD/MM/AA HH:MM
planilha['C3'] = 'INPUT'  # Fornecedor
planilha['F4'] = 'INPUT'  # Ct-e
planilha['F5'] = 'INPUT'  # Motorista
planilha['E16'] = 'INPUT'  # Motivo, vai se iniciar com "motivo:" e concatenar com o real motivo da estadia

valor_celula = planilha['C4'].value  # Lê o valor da célula A1
print(valor_celula)  # Imprime o valor na tela

wb.save('EstadiasCalculadas\Estadia - ' + planilha['F5'] + '.xlsx')  # Salva a planilha com o nome Estadia + o nome do motorista

#####################################################################################################


# Converte xlsx em pdf
# import  aspose.cells
#   from aspose.cells import Workbook
#   workbook = Workbook("input.xlsx")
#   workbook.save("Output.pdf")




import pdfplumber

pdf = pdfplumber.open('estadia/TICKET 1.pdf')
page = pdf.pages[0]
text = page.extract_text()

transportador = text.split('\n')[10].split('-')[1]
numeroNF = int(text.split('\n')[7].split(':')[1])
nomeProduto = text.split('\n')[9].split('-')[1]
pesoNF = text.split('\n')[5].split(": ")[1]
dataHoraSaida = ''

campos = [transportador, numeroNF, nomeProduto, pesoNF, dataHoraSaida]

while True:
    for i in range(4):
        if campos[i] is '':
            input('Campo[' + i + '] = Falta digitar aí vacilão')
    break

print(transportador)
# print(transportador, numeroNF,nomeProduto, pesoNF, dataHoraSaida)
